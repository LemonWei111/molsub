__author__ = "Lemon Wei"
__email__ = "Lemon2922436985@gmail.com"
__version__ = "1.1.0"

DEBUG = False

import os
import cv2
import json
import pickle
import shutil
import random
import pydicom
import argparse
import openpyxl
import numpy as np
import pandas as pd
import nibabel as nib
import SimpleITK as sitk
import matplotlib.pyplot as plt
from PIL import Image
from typing import Dict, Any
from skimage import exposure
from datetime import datetime, date
from radiomics import featureextractor
from matplotlib.colors import ListedColormap
from scipy.ndimage import binary_erosion, binary_dilation, zoom

# 启用详细的CUDA错误报告
# os.environ['CUDA_LAUNCH_BLOCKING'] = '1'
device = 'cpu' # torch.device("cuda" if torch.cuda.is_available() else "cpu")

def get_file_format(file_path):
    # 获取文件扩展名
    root, ext = os.path.splitext(file_path)
    # 检查是否有压缩包内扩展名
    if ext == '.gz':
        second_ext = os.path.splitext(root)[1]
        if second_ext:
            return second_ext[1:].lower()
    else:
        return ext[1:].lower()  # 返回单个扩展名
    
def clahe_img(img, clip_limit=0.003, device = device):
    # 检查输入图像是否为空或全零
    if np.all(img == 0):
        raise ValueError("Input image is empty or all zeros")
    if img.ndim < 2:
        print(f"{img.shape} 图像数组维度不足")
    # 检查图像尺寸
    if img.shape[0] < 8 or img.shape[1] < 8:
        raise ValueError("Input image is too small to apply CLAHE")
    # 自适应直方图均衡化(GPU暂未实现，带来增益也不一定好)
    clahe = exposure.equalize_adapthist(img, clip_limit = clip_limit)
    # 将值缩放到0-255范围
    clahe_img = (clahe * 255).astype(np.uint8)
    return clahe_img

def read_img(img_path, clip_limit = 0.003, device = device):
    img_format = get_file_format(img_path)
    if img_format == 'dcm':
        # 读取.dcm文件
        ds = pydicom.dcmread(img_path)
        # 获取像素数组
        image = ds.pixel_array
        # image = sitk.ReadImage(img_path) # 一维
        # image = np.array(image)
        # TODO. 是否单独增强，逼近潞河数据
    elif img_format == 'jpg':
        # 打开图像
        js = Image.open(img_path)
        # 转换为灰度模式（如果还不是灰度模式）
        if js.mode != 'L':
            js = js.convert('L')
        # 将图像转换为 NumPy 数组
        image = np.array(js)
    else:
        return None
    if DEBUG:
        print(f"Image shape: {image.shape}")
    if clip_limit:
        return clahe_img(image, clip_limit, device = device)
    if DEBUG:
        save_img(image, 'data_process_examples/see_dcm.png')
    return image

def read_nii(nii_path):
    # 读取 .nii 文件
    annotation = nib.load(nii_path)
    # 获取数据数组
    data = annotation.get_fdata()
    # data = sitk.ReadImage(nii_path)
    # data = np.array(data)
    # print(data.shape)
    if DEBUG:
        save_img(data, 'data_process_examples/see_orig_nii.png')
    if DEBUG:
        print(f"Data shape: {data.shape}")
    ## 综合调整：先翻转再旋转
    # nii_data_adjusted = np.flip(data, axis=1)  # 沿x轴翻转
    # nii_data_adjusted = np.rot90(nii_data_adjusted, k=1, axes=(0, 1))  # 逆时针旋转90度
    # 选择一个特定的切片
    # slice_idx = data.shape[2] // 2  # 选择中间切片
    # slice_data = nii_data_adjusted[:, :, slice_idx].squeeze()
    slice_data = data.T[0]
    if DEBUG:
        save_img(slice_data, 'data_process_examples/see_nii.png')
    # 找到白色区域的像素位置
    white_pixels = np.where(slice_data > 0)
    # 计算边界
    if white_pixels[0].size > 0 and white_pixels[1].size > 0:
        top = np.min(white_pixels[0])
        bottom = np.max(white_pixels[0])
        left = np.min(white_pixels[1])
        right = np.max(white_pixels[1])
    else:
        top, bottom, left, right = None, None, None, None
    return top, bottom, left, right, slice_data

def read_json(json_path):
    # 读取 JSON 文件相对路径
    with open(json_path, 'r') as file:
        data = json.load(file)
    # 遍历 shapes 并绘制多边形
    for shape in data['shapes']:
        points = shape['points']
        # 将点列表转换为 (x, y) 坐标对
        x, y = zip(*points)
        # 计算边界框
        left = min(x)
        right = max(x)
        top = min(y)
        bottom = max(y)
    return top, bottom, left, right

def set_bound(annotation_path, bound_size = 0):
    annotation_format = get_file_format(annotation_path)
    if annotation_format == 'nii':
        top, bottom, left, right, cat_label = read_nii(annotation_path)
    if annotation_format == 'json':
        top, bottom, left, right = read_json(annotation_path)
        cat_label = None
    top, bottom, left, right = top - bound_size, bottom + bound_size, left - bound_size, right + bound_size
    return top, bottom, left, right, cat_label

def bound_adjust(x, y, top, bottom, left, right):
    """
    调整边界值，使得最终的区域是一个正方形，并且原来的区域完全包含在新的正方形区域内，
    同时确保新的边界不会超出图像的边界。
    为简化运算，只向上和向左拓展。

    参数:
    top (int): 原始区域的上边界
    bottom (int): 原始区域的下边界
    left (int): 原始区域的左边界
    right (int): 原始区域的右边界
    x (int): 图像的宽度
    y (int): 图像的高度

    返回:
    tuple: (top, bottom, left, right) 调整后的正方形区域的边界
    """
    # 计算原始区域的高度和宽度
    height = bottom - top
    width = right - left
    # 调整边界以形成正方形
    if height > width:
        # 高度大于宽度，向左扩展
        left -= (height - width)
    else:
        # 宽度大于高度，向上扩展
        top -= (width - height)
    # 确保新的边界不会超出图像的边界
    if top < 0:
        bottom -= top
        top = 0
    if left < 0:
        right -= left
        left = 0
    if right > x:
        left -= (right - x)
        right = x
    if bottom > y:
        top -= (bottom - y)
        bottom = y
    return top, bottom, left, right

def get_roi(img_path, annotation_path, clip_limit = 0.003, bound_size = 0, device = device):
    image = read_img(img_path, clip_limit, device = device)
    top, bottom, left, right, cat_label = set_bound(annotation_path, bound_size)
    x, y = image.shape
    top, bottom, left, right = bound_adjust(x, y, top, bottom, left, right)
    if cat_label is not None:
        cat_label = cat_label[top:bottom, left:right]
    return image[top:bottom, left:right], cat_label

def get_roi_v2(img_path, annotation_path, clip_limit = 0.003, bound_size = 0, device = device):
    image = read_img(img_path, clip_limit, device = device)
    top, bottom, left, right, cat_label = set_bound(annotation_path, bound_size)
    x, y = image.shape
    top, bottom, left, right = bound_adjust(x, y, top, bottom, left, right)
    return image, (top, bottom, left, right)

def save_img(image, save_path):
    try:
        plt.imshow(image, cmap='gray')
        plt.axis('off')
        plt.savefig(save_path, bbox_inches='tight', pad_inches=0)  # 保存图像到文件
        plt.close()  # 关闭图形
    except:
        # 分别提取两个通道
        channel_0 = image[:, :, 0]
        channel_1 = image[:, :, 1]
        # 可选：显示每个通道的图像
        plt.figure(figsize=(10, 5))

        plt.subplot(1, 2, 1)
        plt.imshow(channel_0, cmap='gray')
        plt.title('Channel 0')
        plt.axis('off')

        plt.subplot(1, 2, 2)
        plt.imshow(channel_1, cmap='gray')
        plt.title('Channel 1')
        plt.axis('off')

        plt.savefig(save_path, bbox_inches='tight', pad_inches=0)  # 保存图像到文件
        plt.close()  # 关闭图形

def filter_and_traverse_folders(data_dir, all_folders, name):
    # 初始化子文件夹路径列表
    subfolders = []
    # 过滤出以name命名的文件夹
    matching_folders = [f for f in all_folders if f.startswith(name)]
    # 遍历每个匹配的文件夹
    for folder in matching_folders:
        folder_path = os.path.join(data_dir, folder)
        # 获取该文件夹下的所有子文件夹
        subfolders_in_folder = [os.path.join(folder, sf) for sf in os.listdir(folder_path) if os.path.isdir(os.path.join(folder_path, sf))]
        # 将子文件夹路径添加到列表中
        subfolders.extend(subfolders_in_folder)
    if not subfolders:
        subfolders = matching_folders
    return subfolders

def get_roi_through_floader(data_dir, matching_folders, all_folders = None, clip_limit = 0.003, bound_size = 0, device = device):
    data = []
    mask = []
    min_size = float('inf')
    img_paths, annotation_paths = [], []
    for folder in matching_folders:
        folder_path = os.path.join(data_dir, folder)
        # 检查是否已经处理过该文件夹
        if all_folders:
            if all_folders[folder]:
                print(f"警告: 文件夹 {folder} 已经处理过，跳过")
                continue
            all_folders[folder] = 1
        # 获取文件夹内的所有文件
        files = os.listdir(folder_path)
        img_files = []
        anno_files = []
        for f in files:
            ff = get_file_format(f)
            if ff == 'dcm' or ff =='jpg':
                img_files.append(f)
            else:
                anno_files.append(f)
        if len(img_files) != len(anno_files):
            print(f"警告: 文件夹 {folder_path} 内img{len(img_files)}个, anno{len(anno_files)}个, 文件数量不符合要求")
            continue
        for img_file, anno_file in zip(img_files, anno_files):
            # 构建完整的文件路径
            img_path = os.path.join(folder_path, img_file)
            annotation_path = os.path.join(folder_path, anno_file)            
            # 调用get_roi函数处理图像
            roi_data, roi_mask = get_roi(img_path, annotation_path, clip_limit = clip_limit, bound_size = bound_size, device = device)
            min_size = min(min_size, len(roi_data))
            data.append(roi_data)
            mask.append(roi_mask)
            img_paths.append(img_path)
            annotation_paths.append(annotation_path)
    return data, all_folders, min_size, mask, img_paths, annotation_paths

def get_roi_through_floader_v3(data_dir, matching_folders, all_folders = None, img_index = 0, clip_limit = 0.003, device = device):
    img_paths, annotations = [], []
    for folder in matching_folders:
        folder_path = os.path.join(data_dir, folder)
        '''
        # 检查是否已经处理过该文件夹
        if all_folders:
            if all_folders[folder]:
                print(f"警告: 文件夹 {folder} 已经处理过，跳过")
                continue
            all_folders[folder] = 1
        '''
        # 获取文件夹内的所有文件
        files = os.listdir(folder_path)
        img_files = []
        anno_files = []
        for f in files:
            ff = get_file_format(f)
            if ff == 'dcm' or ff =='jpg':
                img_files.append(f)
            else:
                anno_files.append(f)
        if len(img_files) != len(anno_files):
            print(f"警告: 文件夹 {folder_path} 内img{len(img_files)}个, anno{len(anno_files)}个, 文件数量不符合要求")
            continue
        for img_file, anno_file in zip(img_files, anno_files):
            # 构建完整的文件路径
            img_path = os.path.join(folder_path, img_file)
            annotation_path = os.path.join(folder_path, anno_file)            
            # 调用get_roi函数处理图像
            image, annotation = get_roi_v2(img_path, annotation_path, clip_limit = clip_limit, device = device)
            img_name = img_file[:img_file.rfind('.')]
            img_path = f'data/processed/coco/test_luhe/{img_index}_{img_name}.jpg'
            img_index += 1
            # save_img(image, img_path)
            # 保存为JPG，注意这里保持了原DICOM图像的尺寸
            Image.fromarray(image).save(img_path, "JPEG", quality=95)

            img_paths.append(img_path)
            annotations.append(annotation)
    return img_paths, annotations, img_index

def get_roi_through_floader_v4(data_dir, matching_folders, all_folders = None, clip_limit = 0.003, bound_size=100, device = device):
    img_paths = []
    for folder in matching_folders:
        folder_path = os.path.join(data_dir, folder)
        # 检查是否已经处理过该文件夹
        if all_folders:
            if all_folders[folder]:
                print(f"警告: 文件夹 {folder} 已经处理过，跳过")
                continue
            all_folders[folder] = 1
        # 获取文件夹内的所有文件
        files = os.listdir(folder_path)
        img_files = []
        anno_files = []
        for f in files:
            ff = get_file_format(f)
            if ff == 'dcm' or ff =='jpg':
                img_files.append(f)
            else:
                anno_files.append(f)
        if len(img_files) != len(anno_files):
            print(f"警告: 文件夹 {folder_path} 内img{len(img_files)}个, anno{len(anno_files)}个, 文件数量不符合要求")
            continue
        for img_file, anno_file in zip(img_files, anno_files):
            # 构建完整的文件路径
            img_path = os.path.join(folder_path, img_file)
            annotation_path = os.path.join(folder_path, anno_file)            
            # 调用get_roi函数处理图像
            roi, _ = get_roi(img_path, annotation_path, clip_limit = clip_limit, bound_size=bound_size, device = device)
            img_name = img_file[:img_file.rfind('.')]
            img_path = f'data/processed/coco/test/{img_name}.jpg'
            save_img(roi, img_path)
            img_paths.append(img_path)
    return img_paths

def center_pad_array(arr, target_shape):
    """将数组在中心位置填充到指定的形状"""
    pad_h = target_shape - arr.shape[0]
    
    pad_top = pad_h // 2
    pad_bottom = pad_h - pad_top
    
    padding = [(pad_top, pad_bottom), (pad_top, pad_bottom)]
    if DEBUG:
        print(arr.shape, target_shape, padding)
    return np.pad(arr, padding, mode='constant')

def append_data_to_excel(file_path, name, img_path, annotation_path, label):
    """
    将数据递增地存入一个现有的 .xlsx 文件，或者如果文件不存在则创建一个新的文件。

    :param file_path: Excel 文件的路径
    :param data: 要写入的数据，列表形式，每个元素是一个字典
    """
    try:
        # 尝试打开现有文件
        wb = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        # 如果文件不存在，则创建一个新的工作簿
        wb = openpyxl.Workbook()
    
    # 选择活动的工作表
    ws = wb.active

    # 检查是否需要写入表头
    if ws.max_row == 1 and ws.cell(row=1, column=1).value is None:
        headers = ['name', 'img_path', 'annotation_path', 'label']
        # 删除第一行
        ws.delete_rows(1)
        ws.append(headers)

    # 将数据写入工作表
    for img, anno in zip(img_path, annotation_path):
        ws.append([name, img, anno, label])

    # 保存工作簿
    wb.save(file_path)

def process_data(excel_path, data_dir, name_col, label_col, clip_limit = 0.003, bound_size = 0, device = device):
    processed_excel_path = excel_path[:excel_path.rfind('.')] + label_col + '_processed.xlsx'
    if os.path.exists(processed_excel_path):
        os.remove(processed_excel_path)
    # 读取Excel文件
    df = pd.read_excel(excel_path)
    # 初始化数据和标签列表
    data = []
    labels = []
    mask = []
    # 获取data目录下的所有文件夹
    all_folders = {f:0 for f in os.listdir(data_dir) if os.path.isdir(os.path.join(data_dir, f))}
    names = {}

    df_temp = df.copy()
    # 去除空值
    df_temp = df_temp.dropna(subset=[label_col])
    # 过滤出整数值
    df_temp = df_temp[df_temp[label_col].apply(lambda x: isinstance(x, int))]
    # 统计不同值的数量
    num_classes = df_temp[label_col].nunique()
    min_size = float('inf')
    if 'chaoyang' in excel_path:
        # 遍历DataFrame中的每一行
        for _, row in df.iterrows():
            name = str(row[name_col])
            if name not in names.keys():
                names[name] = []
            age = row['年龄']
            names[name].append(age)

            r_value = row[label_col]
            label = r_value - 1 if isinstance(r_value, int) else np.random.randint(0, num_classes - 1) # 缺失默认为随机值
            label = label if label >= 0 else np.random.randint(0, num_classes - 1)

            # 过滤出以name开头的文件夹
            # 正则表达式模式
            # pattern = re.compile(rf'^{re.escape(name)}\d*\s.*$')
            # matching_folders = [os.path.join(data_dir, f) for f in all_folders if pattern.match(f)]
            matching_folders = [f for f in all_folders if f.startswith(name + f' {age}')]
            if not matching_folders:
                see = len(names[name]) - 1
                if see:
                    matching_folders = [f for f in all_folders if f.startswith(name + f'{see} ')]
                else:
                    matching_folders = [f for f in all_folders if f.startswith(name + ' ')]
            if DEBUG:
                print(matching_folders)
            roi_data, all_folders, current_size, roi_mask, img_path, annotation_path = get_roi_through_floader(data_dir, matching_folders, all_folders = all_folders, clip_limit = clip_limit, bound_size = bound_size, device = device)
            min_size = min(min_size, current_size)
            # 将处理后的数据和标签添加到列表中（单个组合）
            data.extend(roi_data)
            mask.extend(roi_mask)
            # 写入工作表
            append_data_to_excel(processed_excel_path, name, img_path, annotation_path, label)
            # label = np.eye(num_classes)[label]
            labels.extend([label] * len(roi_data))
    else:
        # 遍历DataFrame中的每一行
        for _, row in df.iterrows():
            name = str(row[name_col])
            if name not in names.keys():
                names[name] = []
            age = row['年龄']
            names[name].append(age)

            r_value = row[label_col]
            label = r_value - 1 if isinstance(r_value, int) else np.random.randint(0, num_classes - 1) # 缺失默认为随机值
            label = label if label >= 0 else np.random.randint(0, num_classes - 1)

            # 过滤出以name开头的文件夹
            matching_folders = filter_and_traverse_folders(data_dir, all_folders, name)
            '''
            if all_folders[name]:
                print(f"警告: 文件夹 {name} 已经处理过，跳过")
                continue
            all_folders[name] = 1
            '''
            if DEBUG:
                print(matching_folders)
            # 遍历每个匹配的文件夹
            roi_data, _, current_size, roi_mask, img_path, annotation_path = get_roi_through_floader(data_dir, matching_folders, all_folders=None, clip_limit = clip_limit, bound_size = bound_size, device = device)
            min_size = min(min_size, current_size)
            # 将处理后的数据和标签添加到列表中（单个组合）
            data.extend(roi_data)
            mask.extend(roi_mask)
            # 写入工作表
            append_data_to_excel(processed_excel_path, name, img_path, annotation_path, label)
            # label = np.eye(num_classes)[label]
            labels.extend([label] * len(roi_data))
    # 现在data和labels列表包含了所有需要的数据
    print("数据处理完成！")
    return data, labels, num_classes, min_size, mask

def process_data_v2(excel_path, data_dir, name_col, label_col, clip_limit = 0.003, bound_size = 0, device = device):
    processed_excel_path = excel_path[:excel_path.rfind('.')] + label_col + '_processed.xlsx'
    if os.path.exists(processed_excel_path):
        os.remove(processed_excel_path)
    # 读取Excel文件
    df = pd.read_excel(excel_path)
    # 初始化数据和标签列表
    data = []
    labels = []
    mask = None
    # 获取data目录下的所有文件夹
    all_folders = {f:0 for f in os.listdir(data_dir) if os.path.isdir(os.path.join(data_dir, f))}
    names = {}

    df_temp = df.copy()
    # 去除空值
    df_temp = df_temp.dropna(subset=[label_col])
    # 过滤出整数值
    df_temp = df_temp[df_temp[label_col].apply(lambda x: isinstance(x, int))]
    # 统计不同值的数量
    num_classes = df_temp[label_col].nunique()
    min_size = float('inf')
    if 'chaoyang' in excel_path:
        # 遍历DataFrame中的每一行
        for _, row in df.iterrows():
            name = str(row[name_col])
            if name not in names.keys():
                names[name] = []
            age = row['年龄']
            names[name].append(age)

            r_value = row[label_col]
            label = r_value - 1 if isinstance(r_value, int) else np.random.randint(0, num_classes - 1) # 缺失默认为随机值
            label = label if label >= 0 else np.random.randint(0, num_classes - 1)

            # 过滤出以name开头的文件夹
            # 正则表达式模式
            # pattern = re.compile(rf'^{re.escape(name)}\d*\s.*$')
            # matching_folders = [os.path.join(data_dir, f) for f in all_folders if pattern.match(f)]
            matching_folders = [f for f in all_folders if f.startswith(name + f' {age}')]
            if not matching_folders:
                see = len(names[name]) - 1
                if see:
                    matching_folders = [f for f in all_folders if f.startswith(name + f'{see} ')]
                else:
                    matching_folders = [f for f in all_folders if f.startswith(name + ' ')]
            if DEBUG:
                print(matching_folders)
            roi_data, all_folders, current_size, roi_mask, img_path, annotation_path = get_roi_through_floader(data_dir, matching_folders, all_folders = all_folders, clip_limit = clip_limit, bound_size = bound_size, device = device)
            min_size = min(min_size, current_size)
            # 将处理后的数据和标签添加到列表中（单个组合）
            data.extend(roi_data)
            # 写入工作表
            append_data_to_excel(processed_excel_path, name, img_path, annotation_path, label)
            # label = np.eye(num_classes)[label]
            labels.extend([label] * len(roi_data))
    else:
        # 遍历DataFrame中的每一行
        for _, row in df.iterrows():
            name = str(row[name_col])
            if name not in names.keys():
                names[name] = []
            age = row['年龄']
            names[name].append(age)

            r_value = row[label_col]
            label = r_value - 1 if isinstance(r_value, int) else np.random.randint(0, num_classes - 1) # 缺失默认为随机值
            label = label if label >= 0 else np.random.randint(0, num_classes - 1)

            # 过滤出以name开头的文件夹
            matching_folders = filter_and_traverse_folders(data_dir, all_folders, name)
            '''
            if all_folders[name]:
                print(f"警告: 文件夹 {name} 已经处理过，跳过")
                continue
            all_folders[name] = 1
            '''
            if DEBUG:
                print(matching_folders)
            # 遍历每个匹配的文件夹
            roi_data, _, current_size, roi_mask, img_path, annotation_path = get_roi_through_floader(data_dir, matching_folders, all_folders=None, clip_limit = clip_limit, bound_size = bound_size, device = device)
            min_size = min(min_size, current_size)
            # 将处理后的数据和标签添加到列表中（单个组合）
            data.extend(roi_data)
            # 写入工作表
            append_data_to_excel(processed_excel_path, name, img_path, annotation_path, label)
            # label = np.eye(num_classes)[label]
            labels.extend([label] * len(roi_data))
    # 现在data和labels列表包含了所有需要的数据
    print("数据处理完成！")
    return data, labels, num_classes, min_size, mask

def expand_smaller_array(data1, data2):
    """
    将较小的数组放大到与较大的数组相同的尺寸。

    参数:
    - data1: 第一个数组
    - data2: 第二个数组

    返回:
    - 两个尺寸相同的数组
    """
    # 获取两个数组的形状
    shape1 = data1.shape
    shape2 = data2.shape

    # 确定目标形状
    target_shape = (max(shape1[0], shape2[0]), max(shape1[1], shape2[1]))

    # 如果 data1 较小，则放大 data1
    if shape1 != target_shape:
        zoom_factors = (target_shape[0] / shape1[0], target_shape[1] / shape1[1])
        data1 = zoom(data1, zoom_factors, order=1)  # 使用双线性插值

    # 如果 data2 较小，则放大 data2
    if shape2 != target_shape:
        zoom_factors = (target_shape[0] / shape2[0], target_shape[1] / shape2[1])
        data2 = zoom(data2, zoom_factors, order=1)  # 使用双线性插值

    return data1, data2

def get_edge_mask(mask, iterations=4):
    # 计算掩码的边缘 # 增加迭代次数以使边缘更粗
    dilated_mask = binary_dilation(mask, iterations=iterations).astype(int)
    eroded_mask = binary_erosion(mask, iterations=iterations).astype(int)
    edge_mask = dilated_mask - eroded_mask
    return edge_mask

def process_data_from_px(path, args, combine = False, pkl_path=None):
    data = []
    masks = []
    labels = []
    min_size = float('inf')
    load = 0
    clip_limit = args.clip_limit
    bound_size = args.bound_size

    try:
        print(pkl_path)
        print(pkl_path[:pkl_path.rfind('data')] + 'mask' + pkl_path[pkl_path.rfind('data') + len('data'):])
        data, labels, num_classes, min_size = load_data(pkl_path)
        masks = load_mask(pkl_path[:pkl_path.rfind('data')] + 'mask' + pkl_path[pkl_path.rfind('data') + len('data'):])
        print('load from last data')
        load = 1
    except:
        pass

    # 加载数据
    names = []
    # 读取Excel文件
    df = pd.read_excel(path)
    print(df[:3])
    
    df_temp = df.copy()
    # 去除空值
    df_temp = df_temp.dropna(subset=['label'])
    # 过滤出整数值
    df_temp = df_temp[df_temp['label'].apply(lambda x: isinstance(x, int))]
    # 统计不同值的数量
    num_classes = df_temp['label'].nunique()
    # 遍历DataFrame中的每一行
    for index, row in df.iterrows():
        name, img_path, annotation_path, label = row['name'], row['img_path'], row['annotation_path'], row['label']
        names.append(name)
        if not load:
            image = read_img(img_path, clip_limit, device = device)
            _, _, _, _, mask = set_bound(annotation_path, bound_size)

            d, m = get_roi(img_path, annotation_path)
            data.append(d)
            min_size = min(d.shape[0], min_size)
            masks.append(m)
            labels.append(label)

            # img_path = f'test_{name}_{index}.jpg'
            # 保存图像到文件
            # save_path = f'data/processed/luhe/HER2/{label}/' + img_path
            # save_img(d, save_path)

            if DEBUG:
                # 创建一个新的图形
                print(mask)
                # 计算掩码的边缘
                edge_mask = get_edge_mask(mask)
                # 定义一个自定义的颜色映射
                colors = [(0, 0, 0, 0), (1, 0, 0, 1)]  # 0部分透明，1部分浅红色1, 0.8, 0.8, 1
                cmap_mask = ListedColormap(colors)
                fig, axes = plt.subplots(1, 2, figsize=(15, 5))
                # 显示第一个图像
                axes[0].imshow(image, cmap='gray')
                axes[0].imshow(edge_mask, cmap=cmap_mask, vmin=0, vmax=1, interpolation='nearest')
                axes[0].axis('off')
                axes[0].set_title('dcm&nii')
                # 显示第二个图像
                axes[1].imshow(d, cmap='gray')
                axes[1].axis('off')
                axes[1].set_title('ROI')
                plt.savefig(save_path)
                # 关闭图形
                plt.close(fig)    

                one_hot_label = np.eye(2)[label]
                with open('data/train.txt', 'a') as file:
                    # 写入新的一行
                    file.write(f'{img_path} {int(one_hot_label[0])} {int(one_hot_label[1])}\n')
    if combine:
        combine_data = []
        combine_labels = []
        min_size = float('inf')

        i = 0
        while i < len(data) - 1:
            if DEBUG:
                print(names[i], names[i + 1])
            if names[i] == names[i + 1]:
                combine_img = np.stack(expand_smaller_array(data[i], data[i + 1]), axis=-1) # 二通道图
                # save_img(combine_img, 'data_process_examples/new_comine_example.jpg')
                # input('save a combine example')
                combine_data.append(combine_img)
                assert labels[i] == labels[i + 1]
                combine_labels.append(labels[i])
                i += 2
            else:
                combine_img = np.stack([data[i], data[i]], axis=-1) # 二通道图
                combine_data.append(combine_img)
                combine_labels.append(labels[i])
                i += 1
            min_size = min(combine_img.shape[0], min_size)

        if i == len(data) - 1:
            combine_img = np.stack([data[i], data[i]], axis=-1) # 二通道图
            combine_data.append(combine_img)
            combine_labels.append(labels[i])
            min_size = min(combine_img.shape[0], min_size)

        data = combine_data
        labels = combine_labels
    return data, labels, masks, num_classes, min_size

def save_data(data, labels, num_classes, min_size, path = 'data.pkl'):
    # 保存数据
    with open(path, 'wb') as f:
        pickle.dump((data, labels, num_classes, min_size), f)

def load_data(path):
    # 加载数据
    with open(path, 'rb') as f:
        loaded_data, loaded_labels, num_classes, min_size = pickle.load(f)
    return loaded_data, loaded_labels, num_classes, min_size

def save_mask(mask, path = 'mask.pkl'):
    # 保存数据
    with open(path, 'wb') as f:
        pickle.dump(mask, f)

def load_mask(path):
    # 加载数据
    with open(path, 'rb') as f:
        loaded_mask = pickle.load(f)
    return loaded_mask

def unsharp_mask(image, kernel_size=(5, 5), sigma=1.0, amount=1.0, threshold=0):
    """Return a sharpened version of the image, using an unsharp mask."""
    # Convert the image to float32 to avoid overflow
    image = image.astype(np.float32)
    
    # Create a blurred version of the image
    blurred = cv2.GaussianBlur(image, kernel_size, sigma)
    
    # Subtract the blurred image from the original image to get the mask
    mask = image - blurred
    
    # Apply the mask to the original image
    sharpened = image + mask * amount
    
    # Clip the values to the valid range [0, 255]
    sharpened = np.clip(sharpened, 0, 255)
    
    # Convert back to uint8
    sharpened = sharpened.astype(np.uint8)
    
    # Apply a threshold to avoid amplifying noise
    if threshold > 0:
        low_contrast_mask = np.absolute(mask) < threshold
        np.copyto(sharpened, image, where=low_contrast_mask)
    
    return sharpened

def process_rgb_image(image, mask=None):
    # mask
    if mask is not None:
        # image1 = mask # mask作为一个通道
        image1 = image * mask # 肿瘤区域作为一个通道
        # image1 = get_edge_mask(mask) # 肿瘤边界作为一个通道 cv2.error: OpenCV(4.10.0) /croot/opencv-suite_1722029125240/work/modules/imgproc/src/resize.cpp:4030: error: (-215:Assertion failed) func != 0 in function 'resize'
    else:
        # 1. 使用 Otsu 方法生成二值图像 image1
        _, image1 = cv2.threshold(image, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
        # USM
        # image1 = unsharp_mask(image)

    # 2. 使用双边滤波去除原始图像中的尖锐噪声
    image2 = cv2.bilateralFilter(image, d=-1, sigmaColor=20, sigmaSpace=5)

    # 3. 将 image1, image2 与原始图像 image 叠加在一起，生成三通道图像
    # 注意：由于 image1 是二值图像（0 或 255），我们需要将它们转换成 0-1 范围内的浮点数
    # image1 = image1.astype(np.float32) / 255.0

    # 创建三通道图像
    rgb_image = np.stack([image, image1, image2], axis=-1)
    # print(rgb_image.shape)

    if image.dtype != np.uint8:
        rgb_image = (rgb_image * 255).astype(np.uint8)

    if DEBUG:
        # 创建一个新的图形
        fig, axes = plt.subplots(1, 4, figsize=(15, 5))
        # 显示第一个图像
        axes[0].imshow(image, cmap='gray')
        axes[0].axis('off')
        axes[0].set_title('Image 1')
        # 显示第二个图像
        axes[1].imshow(image1, cmap='gray')
        axes[1].axis('off')
        axes[1].set_title('Image 2')
        # 显示第三个图像
        axes[2].imshow(image2, cmap='gray')
        axes[2].axis('off')
        axes[2].set_title('Image 3')
        axes[3].imshow(rgb_image, cmap='gray')
        axes[3].axis('off')
        axes[3].set_title('Image rgb')
        # 保存图像到文件
        save_path = 'data_process_examples/combined_images.png'
        plt.savefig(save_path)
        # 关闭图形
        plt.close(fig)    
    return rgb_image

def generate_mask_from_json(image_path, json_path, output_mask_path='temp.nii.gz'):
    """
    从 JSON 文件和图像文件生成掩码文件。

    参数:
        json_path (str): JSON 文件路径，包含多边形标注信息。
        image_path (str): 图像文件路径，用于获取图像尺寸。
        output_mask_path (str): 输出掩码文件路径。

    返回:
        str: 生成的掩码文件路径。
    """
    # 读取 JSON 文件
    with open(json_path, 'r') as f:
        data = json.load(f)

    # 读取图像文件以获取尺寸
    image = cv2.imread(image_path)
    if image is None:
        raise ValueError(f"无法读取图像文件: {image_path}")
    image_height, image_width = image.shape[:2]

    # 初始化掩码
    mask = np.zeros((image_height, image_width), dtype=np.uint8)

    # 解析多边形标注
    for shape in data['shapes']:
        label = int(shape['label'])  # 假设标签是整数
        points = np.array(shape['points'], dtype=np.int32)
        cv2.fillPoly(mask, [points], label)  # 填充多边形区域

    # 将掩码转换为 SimpleITK 图像
    mask_sitk = sitk.GetImageFromArray(mask)

    # 设置掩码的元数据（例如间距）
    image_sitk = sitk.ReadImage(image_path)
    mask_sitk.CopyInformation(image_sitk)

    # 保存掩码
    sitk.WriteImage(mask_sitk, output_mask_path)

    return output_mask_path

def extract_radiomics(image_paths, mask_paths):
    # 初始化特征提取器
    settings = {
        'binWidth': 25,
        'interpolator': sitk.sitkBSpline,
        'resampledPixelSpacing': None,  # 使用原始像素间距
        'voxelArrayShift': 0,  # 将最小强度值设置为0
    }

    extractor = featureextractor.RadiomicsFeatureExtractor(**settings)
    use_features = ['original_shape_Sphericity', 'original_shape_SurfaceVolumeRatio', 'original_shape_Flatness', 'original_firstorder_Mean', 'original_glcm_Correlation', 'original_firstorder_Skewness']
    features = []

    for image_path, mask_path in zip(image_paths, mask_paths):
        if get_file_format(mask_path) == 'json':
            mask_path = generate_mask_from_json(image_path, mask_path)
        if DEBUG:
            image = sitk.ReadImage(image_path)
            mask = sitk.ReadImage(mask_path)

            # 检查图像和掩码的基本信息
            print("Image Spacing:", image.GetSpacing())
            print("Image Size:", image.GetSize())
            print("Mask Spacing:", mask.GetSpacing())
            print("Mask Size:", mask.GetSize())

            # 确保图像和掩码的尺寸和空间信息一致
            if image.GetSize() != mask.GetSize() or image.GetSpacing() != mask.GetSpacing():
                print("Image and mask dimensions or spacing do not match. Please check the alignment.")
            else:
                print("Image and mask dimensions and spacing match.")
        # 提取特征
        result = extractor.execute(image_path, mask_path)
        if DEBUG:
            print(result)

        # feature = []
        # for f in use_features:
        #     try:
        #         feature.append(np.mean(result[f]))
        #     except:
        #         feature.append(0)
        #     if DEBUG:
        #         print(feature)
        #         input(f'last seen {f}')
        # for chooose
        # feature = [np.mean(result.get(f, 0)) for f in use_features]
        '''
        圆度：original_shape_Sphericity
        凹陷度：original_shape_SurfaceVolumeRatio 和 original_shape_Flatness
        灰度均值：original_firstorder_Mean
        相关性：original_glcm_Correlation
        偏度：original_firstorder_Skewness
        '''
        # for all
        feature = []
        for _, v in result.items():
            if isinstance(v, (int, np.ndarray, np.float64)):
                feature.append(np.mean(v))
        features.append(feature)
    return features

def high_res_image(low_res_image):
    import torch
    from diffusers import StableDiffusionUpscalePipeline

    # 加载预训练的 ESRGAN 模型（下载不下来）
    pipeline = StableDiffusionUpscalePipeline.from_pretrained("model-identifier", torch_dtype=torch.float16)
    # pipeline = StableDiffusionUpscalePipeline.from_pretrained(
    #     "./models/model-identifier",
    #     torch_dtype=torch.float16,
    #     local_files_only=True  # 强制仅使用本地文件，不尝试在线查找
    # )
    pipeline.to("cuda")

    # 进行超分辨率处理
    high_res_image = pipeline(image=low_res_image, num_inference_steps=50).images[0]

    # 保存高分辨率图像
    high_res_image.save('data_process_examples/high_res_image_roi_example_000_CC.png')
    return high_res_image

def find_continuous_regions(arr, value):
    """
    查找数组中连续等于指定值的区域，并返回其索引。
    
    :param arr: 一维数组 (numpy.ndarray)
    :param value: 要查找的值 (int)
    :return: 连续区域的起始和结束索引列表 (list of tuples)
    """
    regions = []
    in_region = False
    start_index = None

    for i, val in enumerate(arr):
        if val == value:
            if not in_region:
                in_region = True
                start_index = i
        else:
            if in_region:
                in_region = False
                regions.append((start_index, i - 1))
        if i == len(arr) - 1 and in_region:
            regions.append((start_index, i))

    return regions

def dilate_border_mask(border_mask, kernel_size=3):
    """
    使用形态学膨胀操作扩展边界的掩码。
    
    :param border_mask: 边界掩码 (numpy.ndarray)
    :param kernel_size: 膨胀内核大小 (int)
    :return: 扩展后的边界掩码 (numpy.ndarray)
    """
    kernel = np.ones((kernel_size, kernel_size), dtype=np.uint8)
    dilated_mask = cv2.dilate(border_mask.astype(np.uint8), kernel, iterations=1)
    return dilated_mask.astype(bool)

def fill_border_with_background_mean(image, mask=None, bound_size=100):
    """
    使用背景均值填充灰度图像的边界（纯0或纯1区域）
    
    :param image: 灰度图像 (numpy.ndarray)
    :param mask: 关注区域的掩码 (numpy.ndarray)，可选
    :return: 填充后的图像 (numpy.ndarray)
    """
    # 确保图像是灰度图像
    if len(image.shape) == 3:
        image = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

    h, w = image.shape[:2]

    # 创建一个与图像大小相同的全1掩码，默认所有像素都属于背景
    background_mask = np.ones_like(image, dtype=bool)

    # 如果提供了mask，则更新background_mask以排除mask中的关注区域
    if mask is not None:
        background_mask[mask > 0] = False

    # 检测边界上的纯0或纯1区域
    border_mask = np.zeros_like(image, dtype=bool)
    
    # 检测行边界上的连续0或255区域
    for i in range(h):
        row = image[i, :]
        for start, end in find_continuous_regions(row, 0) + find_continuous_regions(row, 255):
            if start == 0 or end == w - 1:
                border_mask[i, start:end + 1] = True

    # 检测列边界上的连续0或255区域
    for j in range(w):
        col = image[:, j]
        for start, end in find_continuous_regions(col, 0) + find_continuous_regions(col, 255):
            if start == 0 or end == h - 1:
                border_mask[start:end + 1, j] = True

    # 使用形态学膨胀操作扩展边界的掩码
    border_mask = dilate_border_mask(border_mask, kernel_size=bound_size // 30)

    # 更新background_mask以排除边界区域
    background_mask[border_mask] = False

    # 计算背景区域的平均值
    if np.any(background_mask):
        background_mean = np.mean(image[background_mask])
    else:
        print("警告：没有有效的背景区域用于计算平均值")
        background_mean = 0

    # 用背景平均值填充边界
    filled_image = image.copy()
    filled_image[border_mask] = background_mean
    if DEBUG:
        print(border_mask)
        print(filled_image)
    return filled_image

def convert_excel_to_json(excel_file_path, output_json_path):
    """
    读取指定路径的Excel文件，并将其转换为特定格式的JSON文件。

    参数:
        excel_file_path (str): Excel 文件的路径。
        表头：
        path, top, bottpm, left, right, ER, PR, HER2, ms
        output_json_path (str): 输出 JSON 文件的路径。
    """

    # 读取Excel文件
    df = pd.read_excel(excel_file_path)

    # 创建一个空列表来存储结果
    result_list = {}

    # 遍历DataFrame每一行
    for _, row in df.iterrows():
        image_path = row['path']
        cl2 = ['positive', 'negetive']
        MS = ['Luminal A', 'Luminal B', 'HER2(HR+)', 'HER2(HR-)', 'TN']
        
        # 构建指定格式的字符串，过滤掉可能存在的NaN值
        long_string = f"An image of a breast tumor, top is {row['top']}, bottom is {row['bottom']}, left is {row['left']}, right is {row['right']}, ER {cl2[row['ER']]}, PR {cl2[row['PR']]}, HER2 {cl2[row['HER2']]}, which belongs to {MS[row['ms']]}."
        short_string = f"An image of a {MS[row['ms']]} breast tumor."
        
        # 添加到结果列表中
        result_list[image_path] = [
            long_string,
            short_string
        ]

    # 将列表转换为JSON字符串
    json_data = json.dumps(result_list, ensure_ascii=False, indent=4)

    # 写入JSON文件
    with open(output_json_path, 'w', encoding='utf-8') as json_file:
        json_file.write(json_data)

    print(f"数据已成功写入 {output_json_path}")

def process_data_for_clip(excel_path, data_dir, name_col, output_json_path, clip_limit = 0.003, device = device):
    processed_excel_path = excel_path[:excel_path.rfind('.')] + '_processed.xlsx'
    if os.path.exists(processed_excel_path):
        os.remove(processed_excel_path)
    # 读取Excel文件
    df = pd.read_excel(excel_path)
    # 获取data目录下的所有文件夹
    all_folders = {f:0 for f in os.listdir(data_dir) if os.path.isdir(os.path.join(data_dir, f))}
    names = {}

    try:
        # 尝试打开现有文件
        wb = openpyxl.load_workbook(processed_excel_path)
    except FileNotFoundError:
        # 如果文件不存在，则创建一个新的工作簿
        wb = openpyxl.Workbook()
    
    # 选择活动的工作表
    ws = wb.active

    # 检查是否需要写入表头
    if ws.max_row == 1 and ws.cell(row=1, column=1).value is None:
        headers = ['path', 'top', 'bottom', 'left', 'right', 'ER', 'PR', 'HER2', 'ms']
        # 删除第一行
        ws.delete_rows(1)
        ws.append(headers)

    if 'chaoyang' in excel_path:
        # 遍历DataFrame中的每一行
        for _, row in df.iterrows():
            name = str(row[name_col])
            if name not in names.keys():
                names[name] = []
            age = row['年龄']
            names[name].append(age)

            r_value_er, r_value_pr, r_value_her2, r_value_ms = row['ER'], row['PR'], row['HER2'], row['分子分型']
            label_er = r_value_er - 1 if isinstance(r_value_er, int) else np.random.randint(0, 2) # 缺失默认为随机值
            label_er = label_er if label_er >= 0 else np.random.randint(0, 2)
            label_pr = r_value_pr - 1 if isinstance(r_value_pr, int) else np.random.randint(0, 2) # 缺失默认为随机值
            label_pr = label_pr if label_pr >= 0 else np.random.randint(0, 2)
            label_her2 = r_value_her2 - 1 if isinstance(r_value_her2, int) else np.random.randint(0, 2) # 缺失默认为随机值
            label_her2 = label_her2 if label_her2 >= 0 else np.random.randint(0, 2)
            label_ms = r_value_ms - 1 if isinstance(r_value_ms, int) else np.random.randint(0, 2) # 缺失默认为随机值
            label_ms = label_ms if label_ms >= 0 else np.random.randint(0, 5)

            # 过滤出以name开头的文件夹
            # 正则表达式模式
            # pattern = re.compile(rf'^{re.escape(name)}\d*\s.*$')
            # matching_folders = [os.path.join(data_dir, f) for f in all_folders if pattern.match(f)]
            matching_folders = [f for f in all_folders if f.startswith(name + f' {age}')]
            if not matching_folders:
                see = len(names[name]) - 1
                if see:
                    matching_folders = [f for f in all_folders if f.startswith(name + f'{see} ')]
                else:
                    matching_folders = [f for f in all_folders if f.startswith(name + ' ')]
            if DEBUG:
                print(matching_folders)
            img_paths, annotations, _ = get_roi_through_floader_v3(data_dir, matching_folders, all_folders = all_folders, clip_limit = clip_limit, device = device)
            if DEBUG:
                print(img_paths, annotations)
                input('check')
            for path, anno in zip(img_paths, annotations):
                # 将数据写入工作表
                ws.append([path, anno[0], anno[1], anno[2], anno[3], label_er, label_pr, label_her2, label_ms])
    else:
        # 遍历DataFrame中的每一行
        img_index = 0
        for _, row in df.iterrows():
            name = str(row[name_col])
            if name not in names.keys():
                names[name] = []
            age = row['年龄']
            names[name].append(age)

            r_value_er, r_value_pr, r_value_her2, r_value_ms = row['ER'], row['PR'], row['HER2'], row['分子分型']
            label_er = r_value_er - 1 if isinstance(r_value_er, int) else np.random.randint(0, 2) # 缺失默认为随机值
            label_er = label_er if label_er >= 0 else np.random.randint(0, 2)
            label_pr = r_value_pr - 1 if isinstance(r_value_pr, int) else np.random.randint(0, 2) # 缺失默认为随机值
            label_pr = label_pr if label_pr >= 0 else np.random.randint(0, 2)
            label_her2 = r_value_her2 - 1 if isinstance(r_value_her2, int) else np.random.randint(0, 2) # 缺失默认为随机值
            label_her2 = label_her2 if label_her2 >= 0 else np.random.randint(0, 2)
            label_ms = r_value_ms - 1 if isinstance(r_value_ms, int) else np.random.randint(0, 2) # 缺失默认为随机值
            label_ms = label_ms if label_ms >= 0 else np.random.randint(0, 5)

            # 过滤出以name开头的文件夹
            matching_folders = filter_and_traverse_folders(data_dir, all_folders, name)
            '''
            if all_folders[name]:
                print(f"警告: 文件夹 {name} 已经处理过，跳过")
                continue
            all_folders[name] = 1
            '''
            if DEBUG:
                print(all_folders)
                print(matching_folders)
            # 遍历每个匹配的文件夹
            img_paths, annotations, img_index = get_roi_through_floader_v3(data_dir, matching_folders, all_folders = all_folders, img_index=img_index, clip_limit = clip_limit, device = device)
            for path, anno in zip(img_paths, annotations):
                # 将数据写入工作表
                ws.append([path, anno[0], anno[1], anno[2], anno[3], label_er, label_pr, label_her2, label_ms])
    # 保存工作簿
    wb.save(processed_excel_path)
    # 现在data和labels列表包含了所有需要的数据
    print("数据处理完成！")
    convert_excel_to_json(processed_excel_path, output_json_path)
    convert_excel_to_coco_yolo(processed_excel_path)

def convert_roi_excel_to_json(excel_file_path, output_json_path):
    """
    读取指定路径的Excel文件，并将其转换为特定格式的JSON文件。

    参数:
        excel_file_path (str): Excel 文件的路径。
        表头：
        path, ER, PR, HER2, ms
        output_json_path (str): 输出 JSON 文件的路径。
    """

    # 读取Excel文件
    df = pd.read_excel(excel_file_path)

    # 创建一个空列表来存储结果
    result_list = {}

    # 遍历DataFrame每一行
    for _, row in df.iterrows():
        image_path = row['path']
        cl2 = ['positive', 'negetive']
        MS = ['Luminal A', 'Luminal B', 'HER2(HR+)', 'HER2(HR-)', 'TN']
        
        # 构建指定格式的字符串，过滤掉可能存在的NaN值
        long_string = f"A cropped image of a breast tumor, ER {cl2[row['ER']]}, PR {cl2[row['PR']]}, HER2 {cl2[row['HER2']]}, which belongs to {MS[row['ms']]}."
        short_string = f"A cropped image of a {MS[row['ms']]} breast tumor."
        
        # 添加到结果列表中
        result_list[image_path] = [
            long_string,
            short_string
        ]

    # 将列表转换为JSON字符串
    json_data = json.dumps(result_list, ensure_ascii=False, indent=4)

    # 写入JSON文件
    with open(output_json_path, 'w', encoding='utf-8') as json_file:
        json_file.write(json_data)

    print(f"数据已成功写入 {output_json_path}")

def process_roi_data_for_clip(excel_path, data_dir, name_col, output_json_path, clip_limit=0.003, bound_size=100, device = device):
    processed_excel_path = excel_path[:excel_path.rfind('.')] + '_roi_processed.xlsx'
    if os.path.exists(processed_excel_path):
        os.remove(processed_excel_path)
    # 读取Excel文件
    df = pd.read_excel(excel_path)
    # 获取data目录下的所有文件夹
    all_folders = {f:0 for f in os.listdir(data_dir) if os.path.isdir(os.path.join(data_dir, f))}
    names = {}

    try:
        # 尝试打开现有文件
        wb = openpyxl.load_workbook(processed_excel_path)
    except FileNotFoundError:
        # 如果文件不存在，则创建一个新的工作簿
        wb = openpyxl.Workbook()
    
    # 选择活动的工作表
    ws = wb.active

    # 检查是否需要写入表头
    if ws.max_row == 1 and ws.cell(row=1, column=1).value is None:
        headers = ['path', 'ER', 'PR', 'HER2', 'ms']
        # 删除第一行
        ws.delete_rows(1)
        ws.append(headers)

    if 'chaoyang' in excel_path:
        # 遍历DataFrame中的每一行
        for _, row in df.iterrows():
            name = str(row[name_col])
            if name not in names.keys():
                names[name] = []
            age = row['年龄']
            names[name].append(age)

            r_value_er, r_value_pr, r_value_her2, r_value_ms = row['ER'], row['PR'], row['HER2'], row['分子分型']
            label_er = r_value_er - 1 if isinstance(r_value_er, int) else np.random.randint(0, 2) # 缺失默认为随机值
            label_er = label_er if label_er >= 0 else np.random.randint(0, 2)
            label_pr = r_value_pr - 1 if isinstance(r_value_pr, int) else np.random.randint(0, 2) # 缺失默认为随机值
            label_pr = label_pr if label_pr >= 0 else np.random.randint(0, 2)
            label_her2 = r_value_her2 - 1 if isinstance(r_value_her2, int) else np.random.randint(0, 2) # 缺失默认为随机值
            label_her2 = label_her2 if label_her2 >= 0 else np.random.randint(0, 2)
            label_ms = r_value_ms - 1 if isinstance(r_value_ms, int) else np.random.randint(0, 2) # 缺失默认为随机值
            label_ms = label_ms if label_ms >= 0 else np.random.randint(0, 5)

            # 过滤出以name开头的文件夹
            # 正则表达式模式
            # pattern = re.compile(rf'^{re.escape(name)}\d*\s.*$')
            # matching_folders = [os.path.join(data_dir, f) for f in all_folders if pattern.match(f)]
            matching_folders = [f for f in all_folders if f.startswith(name + f' {age}')]
            if not matching_folders:
                see = len(names[name]) - 1
                if see:
                    matching_folders = [f for f in all_folders if f.startswith(name + f'{see} ')]
                else:
                    matching_folders = [f for f in all_folders if f.startswith(name + ' ')]
            if DEBUG:
                print(matching_folders)
            img_paths = get_roi_through_floader_v4(data_dir, matching_folders, all_folders = all_folders, clip_limit = clip_limit, bound_size=bound_size, device = device)
            if DEBUG:
                print(img_paths)
                input('check')
            for path in img_paths:
                # 将数据写入工作表
                ws.append([path, label_er, label_pr, label_her2, label_ms])
    else:
        # 遍历DataFrame中的每一行
        for _, row in df.iterrows():
            name = str(row[name_col])
            if name not in names.keys():
                names[name] = []
            age = row['年龄']
            names[name].append(age)

            r_value_er, r_value_pr, r_value_her2, r_value_ms = row['ER'], row['PR'], row['HER2'], row['分子分型']
            label_er = r_value_er - 1 if isinstance(r_value_er, int) else np.random.randint(0, 2) # 缺失默认为随机值
            label_er = label_er if label_er >= 0 else np.random.randint(0, 2)
            label_pr = r_value_pr - 1 if isinstance(r_value_pr, int) else np.random.randint(0, 2) # 缺失默认为随机值
            label_pr = label_pr if label_pr >= 0 else np.random.randint(0, 2)
            label_her2 = r_value_her2 - 1 if isinstance(r_value_her2, int) else np.random.randint(0, 2) # 缺失默认为随机值
            label_her2 = label_her2 if label_her2 >= 0 else np.random.randint(0, 2)
            label_ms = r_value_ms - 1 if isinstance(r_value_ms, int) else np.random.randint(0, 2) # 缺失默认为随机值
            label_ms = label_ms if label_ms >= 0 else np.random.randint(0, 5)

            # 过滤出以name开头的文件夹
            matching_folders = filter_and_traverse_folders(data_dir, all_folders, name)
            '''
            if all_folders[name]:
                print(f"警告: 文件夹 {name} 已经处理过，跳过")
                continue
            all_folders[name] = 1
            '''
            if DEBUG:
                print(matching_folders)
            # 遍历每个匹配的文件夹
            img_paths = get_roi_through_floader_v4(data_dir, matching_folders, all_folders = all_folders, clip_limit = clip_limit, bound_size=bound_size, device = device)
            for path in img_paths:
                # 将数据写入工作表
                ws.append([path, label_er, label_pr, label_her2, label_ms])
    # 保存工作簿
    wb.save(processed_excel_path)
    # 现在data和labels列表包含了所有需要的数据
    print("数据处理完成！")
    convert_roi_excel_to_json(processed_excel_path, output_json_path)

def get_image_size(img_path):
    """根据图像路径获取图像尺寸"""
    with Image.open(img_path) as img:
        return img.size # 返回的是(width, height)

def convert_excel_to_coco_yolo(excel_path):
    df = pd.read_excel(excel_path)
    
    coco_format = {
        "images": [],
        "annotations": [],
        "categories": []
    }
    
    yolo_annotations = {}
    
    category_map = {}
    category_id_counter = 0
    
    for index, row in df.iterrows():
        # 确保每个类别都有一个唯一的ID
        if row['ms'] not in category_map:
            category_map[row['ms']] = category_id_counter
            coco_format["categories"].append({"id": category_id_counter, "name": row['ms']})
            category_id_counter += 1
        
        category_id = category_map[row['ms']]
        
        # 根据img_path获取图像尺寸
        width_img, height_img = get_image_size(row['path'])
        
        # 图像信息
        image_info = {"file_name": row['path'], "id": index, "width": width_img, "height": height_img}
        if image_info not in coco_format["images"]:
            coco_format["images"].append(image_info)
        
        # COCO 格式的边界框 [x,y,width,height]
        bbox_coco = [row['left'], row['top'], row['right'] - row['left'], row['bottom'] - row['top']]
        coco_format["annotations"].append({"image_id": index, "bbox": bbox_coco, "category_id": category_id, "id": index})
        
        # YOLO 格式
        center_x = (row['left'] + row['right']) / 2.0
        center_y = (row['top'] + row['bottom']) / 2.0
        width = row['right'] - row['left']
        height = row['bottom'] - row['top']
        
        yolo_format = f"{category_id} {center_x/width_img} {center_y/height_img} {width/width_img} {height/height_img}"
        
        if row['path'] not in yolo_annotations:
            yolo_annotations[row['path']] = []
        yolo_annotations[row['path']].append(yolo_format)

    # 将COCO格式保存为json文件
    with open('data/processed/test_coco_format.json', 'w') as fp:
        json.dump(coco_format, fp)
    
    # 将YOLO格式保存为txt文件，每个图像对应一个txt文件
    for img_path, annotations in yolo_annotations.items():
        filename = img_path.replace('.jpg','.txt').replace('.jpeg','.txt').replace('.png','.txt')
        with open(filename, 'w') as fp:
            for annotation in annotations:
                fp.write(annotation + '\n')

def split_coco_dataset(coco_json_path: str, train_output_path: str, val_output_path: str, split_ratio: float = 0.8):
    """
    将COCO格式的数据集按照给定比例随机划分为训练集和验证集。
    
    参数:
    coco_json_path (str): 输入的完整COCO格式JSON文件路径。
    train_output_path (str): 输出划分后训练集JSON文件的路径。
    val_output_path (str): 输出划分后验证集JSON文件的路径。
    split_ratio (float): 训练集占总数据集的比例，默认是0.8。
    """
    # 加载你的COCO格式JSON文件
    with open(coco_json_path, 'r') as f:
        data: Dict[str, Any] = json.load(f)

    images = data['images']
    annotations = data['annotations']

    # 打乱图像列表
    random.shuffle(images)

    # 计算分割点
    split_point = int(len(images) * split_ratio)

    # 分割图像列表为训练集和验证集
    train_images = images[:split_point]
    val_images = images[split_point:]

    # 创建训练集和验证集的annotation列表
    def filter_annotations(image_ids, annotations):
        return [ann for ann in annotations if ann['image_id'] in image_ids]

    train_image_ids = {img['id'] for img in train_images}
    val_image_ids = {img['id'] for img in val_images}

    train_annotations = filter_annotations(train_image_ids, annotations)
    val_annotations = filter_annotations(val_image_ids, annotations)

    # 构建新的COCO格式字典，同时保留原始JSON文件中的所有信息
    train_data = {
        **{k: v for k, v in data.items() if k not in ['images', 'annotations']},
        'images': train_images,
        'annotations': train_annotations,
    }
    val_data = {
        **{k: v for k, v in data.items() if k not in ['images', 'annotations']},
        'images': val_images,
        'annotations': val_annotations,
    }

    # 将划分后的数据保存到新的JSON文件中
    with open(train_output_path, 'w') as f:
        json.dump(train_data, f)
    with open(val_output_path, 'w') as f:
        json.dump(val_data, f)

    print("数据集已成功分为训练集和验证集，并且保持了COCO格式")

def move_images_and_labels_based_on_json(train_json_path: str, val_json_path: str, images_dir: str = './'):
    """
    根据提供的train和val JSON文件中的image ID信息，
    将对应目录下的图片及其对应的txt标注文件移动到'train/'或'val/'目录。
    
    参数:
    train_json_path (str): 训练集JSON文件路径。
    val_json_path (str): 验证集JSON文件路径。
    images_dir (str): 图像所在的源目录，默认是当前目录。
    """
    # 加载训练集和验证集JSON文件
    with open(train_json_path, 'r') as f:
        train_data = json.load(f)
    with open(val_json_path, 'r') as f:
        val_data = json.load(f)

    # 创建训练集和验证集目录
    train_dir = os.path.join(images_dir, 'train')
    val_dir = os.path.join(images_dir, 'val')
    os.makedirs(train_dir, exist_ok=True)
    os.makedirs(val_dir, exist_ok=True)

    # 收集训练集和验证集的所有image file_name
    train_image_files = {image['id']: image['file_name'] for image in train_data['images']}
    val_image_files = {image['id']: image['file_name'] for image in val_data['images']}
    # print(train_image_files, val_image_files)

    def move_file_with_label(src_path: str, dst_path: str):
        """移动图像文件，并尝试移动同名的txt标注文件"""
        if os.path.exists(src_path):
            shutil.move(src_path, dst_path)
            print(f"Moved to {dst_path}")

            # 查找并移动对应的txt文件
            txt_src_path = os.path.splitext(src_path)[0] + '.txt'
            txt_dst_path = os.path.splitext(dst_path)[0] + '.txt'
            if os.path.exists(txt_src_path):
                shutil.move(txt_src_path, txt_dst_path)
                print(f"Moved label to {txt_dst_path}")

    # 移动训练集图片及对应的txt标注文件
    for _, file_name in train_image_files.items():
        # print(file_name)
        src_path = os.path.join(images_dir, file_name)
        dst_path = os.path.join(train_dir, file_name)
        print(src_path, dst_path)
        move_file_with_label(src_path, dst_path)

    # 移动验证集图片及对应的txt标注文件
    for _, file_name in val_image_files.items():
        src_path = os.path.join(images_dir, file_name)
        dst_path = os.path.join(val_dir, file_name)
        move_file_with_label(src_path, dst_path)

def move_today_images(source_folder, destination_folder):
    """
    将source_folder中今天修改的图像移动到destination_folder。
    
    :param source_folder: 源文件夹路径，字符串形式。
    :param destination_folder: 目标文件夹路径，字符串形式。
    """
    # 如果新文件夹不存在，则创建它
    if not os.path.exists(destination_folder):
        os.makedirs(destination_folder)

    # 获取今天的日期
    today = date.today()
    
    # 遍历源文件夹中的所有文件
    for item in os.listdir(source_folder):
        file_path = os.path.join(source_folder, item)
        
        # 检查是否为文件（排除文件夹）
        if os.path.isfile(file_path):
            # 获取文件的最后修改日期
            modification_time = datetime.fromtimestamp(os.path.getmtime(file_path)).date()
            
            # 检查文件是否在今天被修改
            if modification_time == today:
                # 移动符合条件的文件到新文件夹
                shutil.move(file_path, os.path.join(destination_folder, item))
                print(f"Moved: {item}")

# No use here
def merge_adjacent_labels(inputs, labels):
    """
    视图组合方案1：合并相邻的相同标签，并对相应的输入进行加和和归一化。

    :param inputs: 模型的输入，形状为 (batch_size, height, width)
    :param labels: 标签，形状为 (batch_size,)
    :return: 新的输入和标签（数组）
    """
    new_inputs = []
    new_labels = []
    i = 0
    while i < len(labels):
        start = i
        end = i + 1
        # 找到所有连续相同的标签
        if end < len(labels) and labels[end] == labels[start]:
            end += 1
        
        input_to_merge = inputs[start:end]
        # 找到最大形状
        max_shape = max([x.shape[0] for x in input_to_merge]) # 之前已经处理为方形了
        # 填充所有数组到最大形状
        padded_inputs = [center_pad_array(x, max_shape) for x in input_to_merge]
        
        # 对这些连续的输入进行加和
        merged_input = np.sum(padded_inputs, axis=0) //2

        # 对加和后的输入进行归一化
        # merged_input = merged_input / np.linalg.norm(merged_input)
        
        # 将归一化后的输入和对应的标签添加到新的列表中
        new_inputs.append(merged_input.astype(np.uint8))
        new_labels.append(labels[start])

        # 更新索引i
        i = end
    
    return new_inputs, new_labels

def normalize_image(image):
    """将图像数据归一化到 [0, 1] 范围内"""
    min_val = np.min(image)
    max_val = np.max(image)
    
    if max_val == min_val:
        # 如果最大值和最小值相同，将图像数据设为0
        normalized_image = np.zeros_like(image)
    else:
        normalized_image = (image - min_val) / (max_val - min_val)
    
    return normalized_image

def process_gan_data(img_dir):
    if not img_dir:
        return None
    img_paths = []
    for root, _, files in os.walk(img_dir):
        for filename in files:
            filepath = os.path.join(root, filename)
            img_paths.append(filepath)
    data = [read_img(img_path, clip_limit = 0.003, device = device) for img_path in img_paths]
    return data
   

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description="Process data.")
    parser.add_argument('--data_name', type=int, default=0, help='An integer value, default is 0.')
    parser.add_argument('--bound_size', type=int, default=0, help='An integer value, default is 0.')
    parser.add_argument('--clip_limit', type=float, default=0.003, help='An float value, default is 0.003.')
    parser.add_argument('--label', type=str, default='分子分型', help='A str value, default is 分子分型.')
    parser.add_argument('--train_excel_path', type=str, default='data/mammography subtype dataset/beiyou excel/chaoyang retrospective 233.xlsx', help='A str value.')
    parser.add_argument('--train_data_dir', type=str, default='data/mammography subtype dataset/chaoyang huigu', help='A str value.')

    args = parser.parse_args()
    clip_limit = args.clip_limit
    bound_size = args.bound_size
        
    # HER2 1（阴） 分型 1（LA） L 1 TN 0
    img_path_rc_110 = 'data/mammography subtype dataset/chaoyang huigu/BAILIANDI RCC/ser97311img00002.dcm'
    annotation_path_rc_110 = 'data/mammography subtype dataset/chaoyang huigu/BAILIANDI RCC/1.nii.gz'
    img_path1_rmlo_110 = 'data/mammography subtype dataset/chaoyang huigu/BAILIANDI RMLO/ser97311img00001.dcm'
    annotation_path1_rmlo_110 = 'data/mammography subtype dataset/chaoyang huigu/BAILIANDI RMLO/1.nii.gz'

    save_path_rc_110 = 'data_process_examples/roi_example_110_CC.jpg'
    save_path1_rmlo_110 = 'data_process_examples/roi_example_110_MLO.jpg' 

    # HER2 0（阳） 分型 3（HER2+HR+） L 0 TN 0
    img_path_rc_000 = 'data/mammography subtype dataset/chaoyang huigu/BAOYINHUA RCC/ser121876img00004.dcm'
    annotation_path_rc_000 = 'data/mammography subtype dataset/chaoyang huigu/BAOYINHUA RCC/1.nii.gz'
    img_path1_rmlo_000 = 'data/mammography subtype dataset/chaoyang huigu/BAOYINHUA RMLO/ser121876img00001.dcm'
    annotation_path1_rmlo_000 = 'data/mammography subtype dataset/chaoyang huigu/BAOYINHUA RMLO/1.nii.gz'

    save_path_rc_000 = 'data_process_examples/roi_example_000_CC.jpg'
    save_path1_rmlo_000 = 'data_process_examples/roi_example_000_MLO.jpg' 
    
    # HER2 0（阳） 分型 5 TN L 0 TN 1
    img_path_lc_001 = 'data/mammography subtype dataset/chaoyang huigu/CHENGYUEYING LCC/ser148248img00001.dcm'
    annotation_path_lc_001 = 'data/mammography subtype dataset/chaoyang huigu/CHENGYUEYING LCC/1.nii.gz'
    img_path1_lmlo_001 = 'data/mammography subtype dataset/chaoyang huigu/CHENGYUEYING LMLO/ser148248img00003.dcm'
    annotation_path1_lmlo_001 = 'data/mammography subtype dataset/chaoyang huigu/CHENGYUEYING LMLO/1.nii.gz'

    save_path_lc_001 = 'data_process_examples/roi_example_001_CC.jpg'
    save_path1_lmlo_001 = 'data_process_examples/roi_example_001_MLO.jpg' 
    
    # luhe
    img_path2 = 'data/mammography subtype dataset/luhe mark jpg 323/326188/RCC/2.16.840.1.113669.632.20.20160816.100634573.200195.219.jpg'
    annotation_path2 = 'data/mammography subtype dataset/luhe mark jpg 323/326188/RCC/2.16.840.1.113669.632.20.20160816.100634573.200195.219.json'
    save_path2 = 'data_process_examples/roi_example2.jpg'
    
    
    label_col = args.label
    label = 'ms' if label_col == '分子分型' else label_col
    name_col1 = '姓名'
    
    train_excel_path = args.train_excel_path
    train_data_dir = args.train_data_dir
    val_excel_path = 'data/mammography subtype dataset/beiyou excel/chaoyang prospective 190.xlsx'
    val_data_dir = 'data/mammography subtype dataset/chaoyang qianzhan 190'

    name_col2 = '病例号'
    test_excel_path = 'data/mammography subtype dataset/beiyou excel/luhe 删减后 323.xls'
    test_data_dir = 'data/mammography subtype dataset/luhe mark jpg 323'

    train_save_path = f'data/processed/train_data_{clip_limit}_{bound_size}_{label}.pkl'
    val_save_path = f'data/processed/val_data_{clip_limit}_{bound_size}_{label}.pkl'
    test_save_path = f'data/processed/test_data_{clip_limit}_{bound_size}_{label}.pkl'

    train_mask_path = f'data/processed/train_mask_{clip_limit}_{bound_size}_{label}.pkl'
    val_mask_path = f'data/processed/val_mask_{clip_limit}_{bound_size}_{label}.pkl'
    test_mask_path = f'data/processed/test_mask_{clip_limit}_{bound_size}_{label}.pkl'

    if DEBUG:
        # process_data_for_clip(test_excel_path, test_data_dir, name_col2, 'data/processed/coco/test_luhe_data.json')
        convert_excel_to_coco_yolo('data/mammography subtype dataset/beiyou excel/luhe 删减后 323_processed.xlsx')
        input('chcek')

        split_coco_dataset('data/processed/coco_format.json', 'data/processed/coco/annotations/instance_train2017.json', 'data/processed/coco/annotations/instance_val2017.json')
        move_images_and_labels_based_on_json('data/processed/coco/annotations/instance_train2017.json', 'data/processed/coco/annotations/instance_val2017.json', 'data/processed/yolo')

        process_data_for_clip(train_excel_path, train_data_dir, name_col1, 'data/processed/clip/train_data.json')

        convert_excel_to_json('data/mammography subtype dataset/beiyou excel/chaoyang retrospective 233_processed.xlsx', 'data/processed/clip/train.json')
        save_data([np.array([1]),np.array([2, 2]), np.array([3,4,5])], [1, 2, 3], 5, 8, path = 'data_labels.pkl')
        print(load_data('data_labels.pkl'))

        # high_res_image(Image.open('data_process_examples/roi_example_000_CC.jpg'))
        # input('saved')

        image1, mask1 = get_roi(img_path_rc_000, annotation_path_rc_000, clip_limit, bound_size)
        save_img(fill_border_with_background_mean(image1, mask=mask1), 'data_process_examples/roi_example_000_CC_pad.jpg')
        process_rgb_image(image1)
        input('enter')

        image = read_img(img_path_rc_000, clip_limit, device = device)
        save_img(image, 'data_process_examples/CC.png')
        image = read_img(img_path1_rmlo_110, clip_limit, device = device)
        save_img(image, 'data_process_examples/MLO.png')
        input('test ok')

        image, _ = get_roi(img_path_rc_110, annotation_path_rc_110, clip_limit, bound_size)
        save_img(image, save_path_rc_110)
        image1, mask1 = get_roi(img_path1_rmlo_110, annotation_path1_rmlo_110, clip_limit, bound_size)
        save_img(image1, save_path1_rmlo_110)

        image, _ = get_roi(img_path_rc_000, annotation_path_rc_000, clip_limit, bound_size)
        save_img(image, save_path_rc_000)
        image1, mask1 = get_roi(img_path1_rmlo_000, annotation_path1_rmlo_000, clip_limit, bound_size)
        save_img(image1, save_path1_rmlo_000)

        image, _ = get_roi(img_path_lc_001, annotation_path_lc_001, clip_limit, bound_size)
        save_img(image, save_path_lc_001)
        image1, mask1 = get_roi(img_path1_lmlo_001, annotation_path1_lmlo_001, clip_limit, bound_size)
        save_img(image1, save_path1_lmlo_001)
        input('saved')

        image2, _ = get_roi(img_path2, annotation_path2, clip_limit, bound_size)
        save_img(image2, save_path2)
        
        cv2.imwrite('data_process_examples/roi_example_ostu_rgb.jpg', process_rgb_image(image1))
        cv2.imwrite('data_process_examples/roi_example_mask_rgb.jpg', process_rgb_image(image1, mask1))

        extract_radiomics([img_path1_rmlo_110, img_path2], [annotation_path1_rmlo_110, annotation_path2])
    
        input("img test finish")

    if args.data_name == 0:
        print(train_excel_path)
        train_data, train_labels, num_classes, min_size, train_mask = process_data(train_excel_path, train_data_dir, name_col1, label_col, clip_limit = clip_limit, bound_size = bound_size, device = device)
        print(f"train处理后的数据数量: {len(train_data)}")
        print(f"train处理后的标签数量: {len(train_labels)}")
        print(f"train处理后的最小区域大小: {min_size}")
        save_data(train_data, train_labels, num_classes, min_size, path = train_save_path)
        save_mask(train_mask, path = train_mask_path)

        train_data, train_labels, num_classes, min_size = load_data(train_save_path)
        print(f"train加载的数据数量: {len(train_data)}")
        print(f"train加载的标签数量: {len(train_labels)}")
        print(f"train加载的最小区域大小: {min_size}")
        del train_data, train_labels

    elif args.data_name == 1:
        
        val_data, val_labels, num_classes, min_size, val_mask = process_data(val_excel_path, val_data_dir, name_col1, label_col, clip_limit = clip_limit, bound_size = bound_size, device = device)
        print(f"val处理后的数据数量: {len(val_data)}")
        print(f"val处理后的标签数量: {len(val_labels)}")
        print(f"val处理后的最小区域大小: {min_size}")
        save_data(val_data, val_labels, num_classes, min_size, path = val_save_path)
        save_mask(val_mask, path = val_mask_path)

        val_data, val_labels, num_classes, min_size = load_data(val_save_path)
        print(f"val加载的数据数量: {len(val_data)}")
        print(f"val加载的标签数量: {len(val_labels)}")
        print(f"val加载的最小区域大小: {min_size}")
        del val_data, val_labels
    elif args.data_name == 2:
        
        test_data, test_labels, num_classes, min_size, test_mask = process_data(test_excel_path, test_data_dir, name_col2, label_col, clip_limit = clip_limit, bound_size = bound_size, device = device)
        print(f"test处理后的数据数量: {len(test_data)}")
        print(f"test处理后的标签数量: {len(test_labels)}")
        print(f"test处理后的最小区域大小: {min_size}")
        save_data(test_data, test_labels, num_classes, min_size, path = test_save_path)
        save_mask(test_mask, path = test_mask_path)

        test_data, test_labels, num_classes, min_size = load_data(test_save_path)
        print(f"test加载的数据数量: {len(test_data)}")
        print(f"test加载的标签数量: {len(test_labels)}")
        print(f"test加载的最小区域大小: {min_size}")
        del test_data, test_labels
    elif args.data_name == 3:
        processed_train_excel_path = train_excel_path if train_excel_path.split('.')[0].endswith('_processed') else train_excel_path[:train_excel_path.rfind('.')] + label_col + '_processed.xlsx'
        data, labels, masks, num_classes, min_size = process_data_from_px(processed_train_excel_path, args, False, train_save_path)
        save_data(data, labels, num_classes, min_size, path = f'data/processed/train_data_{clip_limit}_{bound_size}_{label}.pkl')
        # save_mask(masks, path = f'data/processed/train_mask_{clip_limit}_{bound_size}_{label}.pkl')
        train_data, train_labels, num_classes, min_size = load_data(f'data/processed/train_data_{clip_limit}_{bound_size}_{label}.pkl')
        print(f"train加载的数据数量: {len(train_data)}")
        for data in train_data[:3]:
            print(f"数据形状：{data.shape}")
        print(f"train加载的标签数量: {len(train_labels)}")
        print(f"train加载的最小区域大小: {min_size}")
        '''
        # combine from exsit data
        data, labels, masks, num_classes, min_size = process_data_from_px(val_excel_path[:val_excel_path.rfind('.')] + label_col + '_processed.xlsx', args, True, val_save_path)
        save_data(data, labels, num_classes, min_size, path = f'data/processed/val_data_{clip_limit}_{bound_size}_{label}_combine.pkl')
        save_mask(masks, path = f'data/processed/val_mask_{clip_limit}_{bound_size}_{label}.pkl')
        val_data, val_labels, num_classes, min_size = load_data(f'data/processed/val_data_{clip_limit}_{bound_size}_{label}_combine.pkl')
        print(f"合并加载的数据数量: {len(val_data)}")
        for data in val_data[:3]:
            print(f"数据形状：{data.shape}")
        print(f"合并加载的标签数量: {len(val_labels)}")
        print(f"合并加载的最小区域大小: {min_size}")
        
        data, labels, masks, num_classes, min_size = process_data_from_px(test_excel_path[:test_excel_path.rfind('.')] + label_col + '_processed.xlsx', args, True, test_save_path)
        save_data(data, labels, num_classes, min_size, path = f'data/processed/test_data_{clip_limit}_{bound_size}_{label}_combine.pkl')
        save_mask(masks, path = f'data/processed/test_mask_{clip_limit}_{bound_size}_{label}.pkl')
        test_data, test_labels, num_classes, min_size = load_data(f'data/processed/test_data_{clip_limit}_{bound_size}_{label}_combine.pkl')
        print(f"合并加载的数据数量: {len(test_data)}")
        for data in test_data[:3]:
            print(f"数据形状：{data.shape}")
        print(f"合并加载的标签数量: {len(test_labels)}")
        print(f"合并加载的最小区域大小: {min_size}")
        '''
    elif args.data_name == 4:
        train_save_path = f'data/processed/train_img_roi_{clip_limit}_{bound_size}_{label}.pkl'
        val_save_path = f'data/processed/val_img_roi_{clip_limit}_{bound_size}_{label}.pkl'
        test_save_path = f'data/processed/test_img_roi_{clip_limit}_{bound_size}_{label}.pkl'
        '''
        print(train_excel_path)
        
        train_data, train_labels, num_classes, min_size, train_mask = process_data_v2(train_excel_path, train_data_dir, name_col1, label_col, clip_limit = clip_limit, bound_size = bound_size, device = device)
        print(f"train处理后的数据数量: {len(train_data)}")
        print(f"train处理后的标签数量: {len(train_labels)}")
        print(f"train处理后的最小区域大小: {min_size}")
        save_data(train_data, train_labels, num_classes, min_size, path = train_save_path)
        
        train_data, train_labels, num_classes, min_size = load_data(train_save_path)
        print(f"train加载的数据数量: {len(train_data)}")
        print(f"train加载的标签数量: {len(train_labels)}")
        print(f"train加载的最小区域大小: {min_size}")
        print(train_data[0][0].shape, train_data[0][1])
        del train_data, train_labels
     
        val_data, val_labels, num_classes, min_size, val_mask = process_data_v2(val_excel_path, val_data_dir, name_col1, label_col, clip_limit = clip_limit, bound_size = bound_size, device = device)
        print(f"val处理后的数据数量: {len(val_data)}")
        print(f"val处理后的标签数量: {len(val_labels)}")
        print(f"val处理后的最小区域大小: {min_size}")
        save_data(val_data, val_labels, num_classes, min_size, path = val_save_path)

        val_data, val_labels, num_classes, min_size = load_data(val_save_path)
        print(f"val加载的数据数量: {len(val_data)}")
        print(f"val加载的标签数量: {len(val_labels)}")
        print(f"val加载的最小区域大小: {min_size}")
        del val_data, val_labels
        '''
        test_data, test_labels, num_classes, min_size, test_mask = process_data_v2(test_excel_path, test_data_dir, name_col2, label_col, clip_limit = clip_limit, bound_size = bound_size, device = device)
        print(f"test处理后的数据数量: {len(test_data)}")
        print(f"test处理后的标签数量: {len(test_labels)}")
        print(f"test处理后的最小区域大小: {min_size}")
        save_data(test_data, test_labels, num_classes, min_size, path = test_save_path)
        
        test_data, test_labels, num_classes, min_size = load_data(test_save_path)
        print(f"test加载的数据数量: {len(test_data)}")
        print(f"test加载的标签数量: {len(test_labels)}")
        print(f"test加载的最小区域大小: {min_size}")
        del test_data, test_labels
        